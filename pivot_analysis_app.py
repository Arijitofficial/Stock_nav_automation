import pandas as pd
from datetime import datetime, timedelta
import os
import tkinter as tk
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
from dateutil.relativedelta import relativedelta


class PivotAnalysisApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pivot Analysis")
        self.root.geometry("1400x800")
        
        # Data storage
        self.drill_down_df = None
        self.sell_purchase_df = None
        self.current_pivot_data = {}
        self.all_brokers = []
        self.duration_dates = {}
        
        # Load data
        self.load_data()
        
        # Create UI
        self.create_ui()
    
    def load_data(self):
        """Load drill down and sell/purchase tracking data"""
        drill_down_file = './Excels/drill_down_track.csv'
        sell_purchase_file = './Excels/sell_purchase_track.csv'
        
        if os.path.exists(drill_down_file):
            self.drill_down_df = pd.read_csv(drill_down_file)
            self.drill_down_df['date'] = pd.to_datetime(self.drill_down_df['date'])
            self.all_brokers = sorted(self.drill_down_df['broker'].unique().tolist())
        else:
            messagebox.showerror("Error", f"Drill down file not found: {drill_down_file}")
            self.root.destroy()
            return
        
        if os.path.exists(sell_purchase_file):
            self.sell_purchase_df = pd.read_csv(sell_purchase_file)
            self.sell_purchase_df['Purchase Date'] = pd.to_datetime(
                self.sell_purchase_df['Purchase Date'], errors='coerce'
            )
            self.sell_purchase_df['Sell Date'] = pd.to_datetime(
                self.sell_purchase_df['Sell Date'], errors='coerce'
            )
        else:
            messagebox.showwarning("Warning", "Sell/Purchase track file not found. Purchase/Sell values may be inaccurate.")
            self.sell_purchase_df = pd.DataFrame()
    
    def create_ui(self):
        """Create the main user interface"""
        # Create main container
        main_container = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Left panel for controls
        left_panel = ttk.Frame(main_container, width=350)
        main_container.add(left_panel, weight=0)
        
        # Right panel for table display
        right_panel = ttk.Frame(main_container)
        main_container.add(right_panel, weight=1)
        
        # Build left panel
        self.create_control_panel(left_panel)
        
        # Build right panel
        self.create_table_panel(right_panel)
    
    def create_control_panel(self, parent):
        """Create the left control panel"""
        # Title
        title_label = ttk.Label(parent, text="Pivot Analysis", font=('Arial', 14, 'bold'))
        title_label.pack(pady=10)
        
        # Date Selection Frame
        date_frame = ttk.LabelFrame(parent, text="Select End Date", padding=10)
        date_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.end_date_entry = DateEntry(
            date_frame, 
            width=20, 
            background='darkblue',
            foreground='white', 
            date_pattern='yyyy-MM-dd'
        )
        self.end_date_entry.pack(pady=5)
        self.end_date_entry.bind('<<DateEntrySelected>>', self.on_end_date_change)
        
        # Label to show adjusted end date
        self.end_date_label = ttk.Label(date_frame, text="", foreground='blue')
        self.end_date_label.pack(pady=2)
        
        # Duration Selection Frame
        duration_frame = ttk.LabelFrame(parent, text="Duration Start Dates", padding=10)
        duration_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.duration_widgets = {}
        durations = [('1M', 1), ('3M', 3), ('6M', 6), ('9M', 9), ('12M', 12)]
        
        for duration_code, months in durations:
            row_frame = ttk.Frame(duration_frame)
            row_frame.pack(fill=tk.X, pady=3)
            
            ttk.Label(row_frame, text=duration_code, width=5).pack(side=tk.LEFT, padx=5)
            
            date_entry = ttk.Entry(row_frame, width=12)
            date_entry.pack(side=tk.LEFT, padx=5)
            date_entry.bind('<FocusOut>', lambda e, dc=duration_code: self.on_duration_date_change(dc))
            date_entry.bind('<Return>', lambda e, dc=duration_code: self.on_duration_date_change(dc))
            
            available_label = ttk.Label(row_frame, text="", foreground='blue')
            available_label.pack(side=tk.LEFT, padx=5)
            
            self.duration_widgets[duration_code] = {
                'months': months,
                'entry': date_entry,
                'label': available_label
            }
        
        # Broker Selection Frame
        broker_frame = ttk.LabelFrame(parent, text="Select Broker", padding=10)
        broker_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Tree view for broker selection
        self.tree = ttk.Treeview(broker_frame, selectmode='browse', height=10)
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar for tree
        tree_scroll = ttk.Scrollbar(broker_frame, orient="vertical", command=self.tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        # Populate tree with brokers only
        for broker in self.all_brokers:
            self.tree.insert('', 'end', text=broker, values=(broker,))
        
        # Bind selection event
        self.tree.bind('<<TreeviewSelect>>', self.on_broker_select)
        
        # Buttons Frame
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(
            button_frame,
            text="Generate Pivots For Broker",
            command=self.generate_all_pivots
        ).pack(fill=tk.X, pady=2)
        
        ttk.Button(
            button_frame,
            text="Download Selected Broker",
            command=self.download_selected_broker
        ).pack(fill=tk.X, pady=2)
        
        ttk.Button(
            button_frame,
            text="Download All Brokers",
            command=self.download_all_brokers
        ).pack(fill=tk.X, pady=2)
        
        # ttk.Button(
        #     button_frame,
        #     text="Generate Drill Down CSV",
        #     command=self.generate_drill_down_csv
        # ).pack(fill=tk.X, pady=2)
    
    def on_end_date_change(self, event=None):
        """Handle end date change and calculate duration start dates"""
        try:
            end_date_str = self.end_date_entry.get()
            if not end_date_str or end_date_str.strip() == '':
                return
                
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            # Find and set closest available end date
            closest_end_date = self.find_closest_date(end_date)
            if closest_end_date:
                # Update the DateEntry widget
                self.end_date_entry.set_date(closest_end_date)
                self.end_date_label.config(
                    text=f"Adjusted to: {closest_end_date.strftime('%Y-%m-%d')}"
                )
                end_date = closest_end_date
            else:
                self.end_date_label.config(text="No data available")
                return
            
            # Calculate duration start dates
            for duration_code, widget_info in self.duration_widgets.items():
                months = widget_info['months']
                
                # Calculate target start date (same day, N months earlier)
                target_start_date = end_date - relativedelta(months=months)
                
                # Find closest available date
                available_date = self.find_closest_date(target_start_date)
                
                if available_date:
                    # Set the entry field
                    widget_info['entry'].delete(0, tk.END)
                    widget_info['entry'].insert(0, available_date.strftime('%Y-%m-%d'))
                    
                    # Show available date in label
                    widget_info['label'].config(
                        text=f"(Available: {available_date.strftime('%Y-%m-%d')})"
                    )
                else:
                    widget_info['entry'].delete(0, tk.END)
                    widget_info['entry'].insert(0, target_start_date.strftime('%Y-%m-%d'))
                    widget_info['label'].config(text="(No data)")
                    
        except ValueError:
            # Invalid date format
            pass
        except Exception as e:
            messagebox.showerror("Error", f"Error calculating dates: {str(e)}")
    
    def on_duration_date_change(self, duration_code):
        """Handle manual change to duration start date"""
        try:
            widget_info = self.duration_widgets[duration_code]
            date_str = widget_info['entry'].get()
            
            if not date_str or date_str.strip() == '':
                widget_info['label'].config(text="(Empty)")
                return
            
            # Parse the entered date
            entered_date = datetime.strptime(date_str, '%Y-%m-%d')
            
            # Find closest available date
            closest_date = self.find_closest_date(entered_date)
            
            if closest_date:
                # Update entry with closest date
                widget_info['entry'].delete(0, tk.END)
                widget_info['entry'].insert(0, closest_date.strftime('%Y-%m-%d'))
                
                # Update label
                widget_info['label'].config(
                    text=f"(Available: {closest_date.strftime('%Y-%m-%d')})"
                )
            else:
                widget_info['label'].config(text="(No data)")
                
        except ValueError:
            # Invalid date format
            widget_info = self.duration_widgets[duration_code]
            widget_info['label'].config(text="(Invalid format)")
        except Exception as e:
            print(f"Error processing duration date: {str(e)}")
    
    def find_closest_date(self, target_date):
        """Find the closest available date in the drill_down data"""
        available_dates = self.drill_down_df['date'].unique()
        available_dates = pd.to_datetime(available_dates)
        available_dates = sorted(available_dates)
        
        # Find the closest date that is <= target_date
        valid_dates = [d for d in available_dates if d <= target_date]
        
        if valid_dates:
            return max(valid_dates)
        
        return None
    
    def on_broker_select(self, event):
        """Handle broker selection"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        values = item['values']
        
        if values:
            self.current_broker = values[0]
    
    def create_table_panel(self, parent):
        """Create the right panel for table display"""
        # Title
        self.table_title = ttk.Label(
            parent, 
            text="Generate pivots to view data",
            font=('Arial', 12, 'bold')
        )
        self.table_title.pack(pady=10)
        
        # Notebook for tabs
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create tabs for each duration
        self.table_frames = {}
        durations = ['1M', '3M', '6M', '9M', '12M']
        
        for duration in durations:
            tab_frame = ttk.Frame(self.notebook)
            self.notebook.add(tab_frame, text=duration)
            
            # Create table in this tab
            table_tree = self.create_duration_table(tab_frame)
            self.table_frames[duration] = table_tree
    
    def create_duration_table(self, parent):
        """Create a table for a specific duration"""
        # Table frame
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create Treeview for table
        columns = (
            'Stock Name',
            'Qty (Start)',
            'Value (Start)',
            'Qty (End)',
            'Value (End)',
            'Purchase Value',
            'Sell Value',
            'Total P&L',
            '% P&L'
        )
        
        table_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=30
        )
        
        # Define column headings and widths
        column_widths = {
            'Stock Name': 150,
            'Qty (Start)': 100,
            'Value (Start)': 120,
            'Qty (End)': 100,
            'Value (End)': 120,
            'Purchase Value': 130,
            'Sell Value': 120,
            'Total P&L': 120,
            '% P&L': 100
        }
        
        for col in columns:
            table_tree.heading(col, text=col)
            table_tree.column(col, width=column_widths.get(col, 100), anchor='center')
        
        # Scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=table_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=table_tree.xview)
        table_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Grid layout
        table_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        return table_tree
    
    def generate_all_pivots(self):
        """Generate pivot tables for all durations"""
        if not hasattr(self, 'current_broker'):
            messagebox.showwarning("Warning", "Please select a broker first")
            return
        
        # Validate end date
        try:
            end_date_str = self.end_date_entry.get()
            if not end_date_str or end_date_str.strip() == '':
                messagebox.showwarning("Warning", "Please select an end date")
                return
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Error", "Invalid end date format")
            return
        
        broker = self.current_broker
        self.current_pivot_data[broker] = {}
        
        for duration_code, widget_info in self.duration_widgets.items():
            try:
                start_date_str = widget_info['entry'].get()
                
                if not start_date_str or start_date_str.strip() == '':
                    messagebox.showwarning("Warning", f"Start date for {duration_code} is empty")
                    continue
                    
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                
                # Generate pivot data
                pivot_df, actual_start_date, actual_end_date = self.calculate_pivot_table(
                    broker, start_date, end_date
                )
                
                if pivot_df is not None and not pivot_df.empty and actual_start_date and actual_end_date:
                    # Store pivot data
                    self.current_pivot_data[broker][duration_code] = {
                        'start_date': actual_start_date,
                        'end_date': actual_end_date,
                        'data': pivot_df
                    }
                    
                    # Display in corresponding tab
                    self.display_pivot_table(
                        pivot_df, 
                        broker, 
                        duration_code, 
                        actual_start_date, 
                        actual_end_date,
                        self.table_frames[duration_code]
                    )
                else:
                    # Clear the tab if no data
                    table_tree = self.table_frames[duration_code]
                    for item in table_tree.get_children():
                        table_tree.delete(item)
                    
            except ValueError:
                messagebox.showerror("Error", f"Invalid date format for {duration_code}")
            except Exception as e:
                messagebox.showerror("Error", f"Error generating {duration_code}: {str(e)}")
        
        self.table_title.config(text=f"Pivot Tables for {broker}")
        messagebox.showinfo("Success", f"All pivot tables generated for {broker}")
    

    def add_total_row(self, pivot_df: pd.DataFrame) -> pd.DataFrame:
        """Add TOTAL row directly into pivot dataframe"""

        if pivot_df.empty:
            return pivot_df

        total_start_value = pivot_df['Value (Start)'].sum()
        total_end_value = pivot_df['Value (End)'].sum()
        total_purchase = pivot_df['Purchase Value'].sum()
        total_sell = pivot_df['Sell Value'].sum()
        total_pl = pivot_df['Total P&L'].sum()

        if total_start_value > 0:
            avg_pct_pl = (total_pl / total_start_value) * 100
        else:
            avg_pct_pl = 0

        total_row = pd.DataFrame([{
            'Stock Name': 'TOTAL',
            'Qty (Start)': '',
            'Value (Start)': round(total_start_value, 2),
            'Qty (End)': '',
            'Value (End)': round(total_end_value, 2),
            'Purchase Value': round(total_purchase, 2),
            'Sell Value': round(total_sell, 2),
            'Total P&L': round(total_pl, 2),
            '% P&L': round(avg_pct_pl, 2)
        }])

        return pd.concat([pivot_df, total_row], ignore_index=True)


    def calculate_pivot_table(self, broker, start_date, end_date):
        """Calculate pivot table data with aggregation across files"""
        # Filter drill down data for the broker
        broker_df = self.drill_down_df[self.drill_down_df['broker'] == broker].copy()
        
        if broker_df.empty:
            return None, None, None
        
        # Get unique stocks
        stocks = broker_df['share name'].unique()
        
        pivot_data = []
        actual_start_date = None
        actual_end_date = None
        
        for stock in stocks:
            stock_df = broker_df[broker_df['share name'] == stock]
            
            # Aggregate data at start date (closest date <= start_date) across all files
            start_data = stock_df[stock_df['date'] <= start_date]
            if not start_data.empty:
                # Get the latest date at or before start_date
                latest_start_date = start_data['date'].max()
                start_data_on_date = start_data[start_data['date'] == latest_start_date]
                
                # Track actual start date
                if actual_start_date is None or latest_start_date < actual_start_date:
                    actual_start_date = latest_start_date
                
                # Aggregate across all files for this stock on this date
                qty_start = start_data_on_date['quantity'].sum()
                value_start = start_data_on_date['total market value'].sum()
            else:
                qty_start = 0
                value_start = 0
            
            # Aggregate data at end date (closest date <= end_date) across all files
            end_data = stock_df[stock_df['date'] <= end_date]
            if not end_data.empty:
                # Get the latest date at or before end_date
                latest_end_date = end_data['date'].max()
                end_data_on_date = end_data[end_data['date'] == latest_end_date]
                
                # Track actual end date
                if actual_end_date is None or latest_end_date > actual_end_date:
                    actual_end_date = latest_end_date
                
                # Aggregate across all files for this stock on this date
                qty_end = end_data_on_date['quantity'].sum()
                value_end = end_data_on_date['total market value'].sum()
            else:
                qty_end = 0
                value_end = 0
            
            # Calculate purchase and sell values from sell_purchase_track
            purchase_value = 0
            sell_value = 0
            
            if not self.sell_purchase_df.empty:
                stock_transactions = self.sell_purchase_df[
                    (self.sell_purchase_df['Stock Symbol'] == stock) &
                    (self.sell_purchase_df['Broker'] == broker)
                ]
                
                # Purchases in range - aggregate across all files
                purchases = stock_transactions[
                    (stock_transactions['Purchase Date'] >= start_date) &
                    (stock_transactions['Purchase Date'] <= end_date) &
                    (stock_transactions['Purchase Price'].notna())
                ]
                purchase_value = (purchases['Purchase Price'] * purchases['Quantity']).sum()
                
                # Sells in range - aggregate across all files
                sells = stock_transactions[
                    (stock_transactions['Sell Date'] >= start_date) &
                    (stock_transactions['Sell Date'] <= end_date) &
                    (stock_transactions['Sell Price'].notna())
                ]
                sell_value = (sells['Sell Price'] * sells['Quantity']).sum()
            
            # Calculate net change and percentage
            net_change = value_end - value_start + sell_value - purchase_value
            
            # Calculate percentage P&L
            if value_start > 0:
                pct_pl = (net_change / value_start) * 100
            else:
                pct_pl = 0
            
            # Only include if there's any activity
            if qty_start > 0 or qty_end > 0 or purchase_value > 0 or sell_value > 0:
                pivot_data.append({
                    'Stock Name': stock,
                    'Qty (Start)': qty_start,
                    'Value (Start)': round(value_start, 2),
                    'Qty (End)': qty_end,
                    'Value (End)': round(value_end, 2),
                    'Purchase Value': round(purchase_value, 2),
                    'Sell Value': round(sell_value, 2),
                    'Total P&L': round(net_change, 2),
                    '% P&L': round(pct_pl, 2)
                })

        pivot_df = pd.DataFrame(pivot_data)
        if pivot_df is not None and not pivot_df.empty:
            pivot_df = self.add_total_row(pivot_df)

        return pivot_df, actual_start_date, actual_end_date
    
    def display_pivot_table(self, pivot_df, broker, duration, start_date, end_date, table_tree):
        """Display pivot table in the treeview"""
        # Clear existing data
        for item in table_tree.get_children():
            table_tree.delete(item)
        
        # Insert data
        for _, row in pivot_df.iterrows():
            is_total = row['Stock Name'] == 'TOTAL'

            values = (
                row['Stock Name'],
                row['Qty (Start)'],
                f"₹{row['Value (Start)']:,.2f}" if row['Value (Start)'] != '' else '',
                row['Qty (End)'],
                f"₹{row['Value (End)']:,.2f}" if row['Value (End)'] != '' else '',
                f"₹{row['Purchase Value']:,.2f}",
                f"₹{row['Sell Value']:,.2f}",
                f"₹{row['Total P&L']:,.2f}",
                f"{row['% P&L']:.2f}%"
            )

            tag = ('total',) if is_total else ()
            table_tree.insert('', 'end', values=values, tags=tag)

        table_tree.tag_configure('total', background='lightgray', font=('Arial', 9, 'bold'))

    
    def download_selected_broker(self):
        """Download Excel file with all duration sheets for selected broker"""
        if not hasattr(self, 'current_broker'):
            messagebox.showwarning("Warning", "Please select a broker first")
            return
        
        broker = self.current_broker
        
        if broker not in self.current_pivot_data or not self.current_pivot_data[broker]:
            messagebox.showwarning("Warning", "No pivot data generated. Please generate pivots first.")
            return
        
        # Create output directory
        output_dir = './pivot_tables'
        os.makedirs(output_dir, exist_ok=True)
        
        end_date = datetime.strptime(self.end_date_entry.get(), '%Y-%m-%d')
        filename = f"pivot_{broker}_{end_date.strftime('%Y-%m-%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for duration_code in ['1M', '3M', '6M', '9M', '12M']:
                if duration_code in self.current_pivot_data[broker]:
                    pivot_info = self.current_pivot_data[broker][duration_code]
                    
                    # Format dates
                    start_date = pivot_info['start_date']
                    end_date_actual = pivot_info['end_date']
                    
                    if isinstance(start_date, pd.Timestamp):
                        start_date_str = start_date.strftime('%Y-%m-%d')
                    else:
                        start_date_str = start_date.strftime('%Y-%m-%d')
                        
                    if isinstance(end_date_actual, pd.Timestamp):
                        end_date_str = end_date_actual.strftime('%Y-%m-%d')
                    else:
                        end_date_str = end_date_actual.strftime('%Y-%m-%d')
                    
                    # Rename columns with actual dates
                    df_to_save = pivot_info['data'].copy()
                    print(df_to_save)
                    df_to_save.columns = [
                        'Stock Name',
                        f'Quantity as on {start_date_str}',
                        f'Total Value as on {start_date_str}',
                        f'Quantity as on {end_date_str}',
                        f'Total Value as on {end_date_str}',
                        'Purchase Value',
                        'Sell Value',
                        'Total P&L',
                        '% P&L'
                    ]
                    
                    # Write to Excel sheet
                    df_to_save.to_excel(writer, sheet_name=duration_code, index=False)

                    worksheet = writer.sheets[duration_code]
                    self.highlight_last_row_excel(worksheet)
        
        messagebox.showinfo("Success", f"Excel file saved to:\n{filepath}")
    
    def download_all_brokers(self):
        """Generate and download Excel files for all brokers with all duration sheets"""

        # -------------------- VALIDATE END DATE --------------------
        try:
            end_date_str = self.end_date_entry.get()
            if not end_date_str or end_date_str.strip() == '':
                messagebox.showwarning("Warning", "Please select an end date")
                return
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Error", "Invalid end date format")
            return

        output_dir = './pivot_tables'
        os.makedirs(output_dir, exist_ok=True)

        # -------------------- UI --------------------
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Generating Pivot Tables")
        progress_window.geometry("420x120")
        progress_window.protocol("WM_DELETE_WINDOW", lambda: None)

        progress_label = ttk.Label(progress_window, text="Preparing...")
        progress_label.pack(pady=10)

        progress_bar = ttk.Progressbar(
            progress_window,
            length=380,
            mode='determinate',
            maximum=len(self.all_brokers)
        )
        progress_bar.pack(pady=10)

        # -------------------- TRACKING --------------------
        no_data_brokers = []
        failed_brokers = []

        count = 0

        # -------------------- MAIN LOOP --------------------
        for broker in self.all_brokers:
            filename = f"pivot_{broker}_{end_date.strftime('%Y-%m-%d')}.xlsx"
            filepath = os.path.join(output_dir, filename)

            sheet_written = False

            try:
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    for duration_code, widget_info in self.duration_widgets.items():

                        start_date_str = widget_info['entry'].get()
                        
                        if not start_date_str or start_date_str.strip() == '':
                            continue

                        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')

                        pivot_df, actual_start_date, actual_end_date = self.calculate_pivot_table(
                            broker, start_date, end_date
                        )

                        if pivot_df is None or pivot_df.empty:
                            continue

                        # TOTAL already exists inside pivot_df
                        df_to_save = pivot_df.copy()

                        if actual_start_date is None or actual_end_date is None:
                            continue  # skip this duration safely

                        # Format column headers
                        start_str = actual_start_date.strftime('%Y-%m-%d')
                        end_str = actual_end_date.strftime('%Y-%m-%d')

                        df_to_save.columns = [
                            'Stock Name',
                            f'Quantity as on {start_str}',
                            f'Total Value as on {start_str}',
                            f'Quantity as on {end_str}',
                            f'Total Value as on {end_str}',
                            'Purchase Value',
                            'Sell Value',
                            'Total P&L',
                            '% P&L'
                        ]

                        df_to_save.to_excel(writer, sheet_name=duration_code, index=False)

                        worksheet = writer.sheets[duration_code]
                        self.highlight_last_row_excel(worksheet)

                        sheet_written = True

                if not sheet_written:
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    no_data_brokers.append(broker)

            # -------------------- PERMISSION ERROR --------------------
            except PermissionError:
                progress_window.destroy()
                messagebox.showerror(
                    "File in Use",
                    f"Permission denied while saving:\n\n{filename}\n\n"
                    "Please close any open Excel files with this name,\n"
                    "then close this window and try again."
                )
                return

            # -------------------- OTHER ERRORS --------------------
            except Exception as e:
                if os.path.exists(filepath):
                    os.remove(filepath)
                failed_brokers.append(f"{broker}: {str(e)}")

            # -------------------- UPDATE UI --------------------
            count += 1
            progress_bar['value'] = count
            progress_label.config(
                text=f"Processing {broker} ({count}/{len(self.all_brokers)})"
            )
            progress_window.update_idletasks()

        progress_window.destroy()

        # -------------------- FINAL SUMMARY --------------------
        summary = "Download completed.\n\n"

        if no_data_brokers:
            summary += "No data found for:\n"
            summary += ", ".join(no_data_brokers) + "\n\n"

        if failed_brokers:
            summary += "Failed brokers:\n"
            summary += "\n".join(failed_brokers)

        messagebox.showinfo("Completed", summary)


    def highlight_last_row_excel(self, worksheet):
        from openpyxl.styles import Font, PatternFill, Border, Side
        """Highlight last row (TOTAL) in an openpyxl worksheet"""

        last_row = worksheet.max_row
        last_col = worksheet.max_column

        fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        font = Font(bold=True)
        border = Border(
            top=Side(style="medium"),
            bottom=Side(style="thin")
        )

        for col in range(1, last_col + 1):
            cell = worksheet.cell(row=last_row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = border


    def generate_drill_down_csv(self):
        """Generate filtered drill down CSV (original functionality)"""
        # Create date selection dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Generate Drill Down CSV")
        dialog.geometry("400x250")
        
        ttk.Label(dialog, text="Select Date Range for Drill Down", font=('Arial', 12, 'bold')).pack(pady=10)
        
        ttk.Label(dialog, text="Start Date:").pack(pady=5)
        start_date_entry = DateEntry(
            dialog,
            width=20,
            background='darkblue',
            foreground='white',
            date_pattern='yyyy-MM-dd'
        )
        start_date_entry.pack(pady=5)
        
        ttk.Label(dialog, text="End Date:").pack(pady=5)
        end_date_entry = DateEntry(
            dialog,
            width=20,
            background='darkblue',
            foreground='white',
            date_pattern='yyyy-MM-dd'
        )
        end_date_entry.pack(pady=5)
        
        def on_generate():
            start_date = datetime.strptime(start_date_entry.get(), '%Y-%m-%d')
            end_date = datetime.strptime(end_date_entry.get(), '%Y-%m-%d')
            
            if start_date > end_date:
                messagebox.showerror("Error", "Start date must be before end date")
                return
            
            # Filter data
            df_range = self.drill_down_df[
                (self.drill_down_df['date'] >= start_date) &
                (self.drill_down_df['date'] <= end_date)
            ]
            filtered_df = df_range[
                df_range['date'].isin([df_range['date'].min(), df_range['date'].max()])
            ]
            
            if filtered_df.empty:
                messagebox.showinfo("No Data", "No data found in the given date range")
                return
            
            # Save
            output_dir = './drill_downs'
            os.makedirs(output_dir, exist_ok=True)
            
            output_file = os.path.join(
                output_dir,
                f'drill_down_{start_date.date()}_to_{end_date.date()}.csv'
            )
            filtered_df.to_csv(output_file, index=False)
            
            messagebox.showinfo("Success", f"Drill down CSV saved to:\n{output_file}")
            dialog.destroy()
        
        ttk.Button(dialog, text="Generate", command=on_generate).pack(pady=20)


def main():
    root = tk.Tk()
    app = PivotAnalysisApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()